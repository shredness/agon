from fastapi import FastAPI, HTTPException, Depends, status, Request, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel
from typing import Optional
import json, os, re
import httpx
import pyotp, qrcode, io, base64
from datetime import datetime, timedelta, date as _date
from jose import JWTError, jwt
import bcrypt as _bcrypt

# Import new dual-database module
import db as database_module
from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: initialize database
    database_module.init_db()
    yield
    # Shutdown: (connection pool cleanup handled by db module)

app = FastAPI(title="Agon API", lifespan=lifespan)

ALLOWED_ORIGINS = os.environ.get(
    "ALLOWED_ORIGINS",
    "https://agon.savo.us,http://localhost,http://localhost:3800"
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def db_connection_scope(request: Request, call_next):
    """Open a per-request DB connection scope and reap any connection that wasn't
    released, so an exception between get_db() and conn.close() can't leak a pooled
    connection. Without this, the 10-slot pool drains and the API becomes unreachable."""
    token = database_module.begin_request_scope()
    try:
        return await call_next(request)
    finally:
        database_module.end_request_scope(token)

SECRET_KEY        = os.environ.get("SECRET_KEY", "agon-change-this-in-production")
ALGORITHM         = "HS256"
TOKEN_EXPIRE_DAYS = 30

# Refuse to run with the known default secret unless explicitly allowed (e.g. local dev)
_DEFAULT_SECRET = "agon-change-this-in-production"
if SECRET_KEY == _DEFAULT_SECRET and os.environ.get("ALLOW_DEFAULT_SECRET", "false").lower() != "true":
    raise RuntimeError(
        "SECRET_KEY is still the insecure default. Set a strong SECRET_KEY env var, "
        "or set ALLOW_DEFAULT_SECRET=true for local development only."
    )

oauth2 = OAuth2PasswordBearer(tokenUrl="/auth/login")

ADMIN_USER         = os.environ.get("ADMIN_USER")
ADMIN_PASS         = os.environ.get("ADMIN_PASS")
# These shipped with a weak default in early builds. Never hardcode a fallback —
# and refuse to boot if the old known value is still present in the environment.
if ADMIN_PASS == "pR0m3th3us" and os.environ.get("ALLOW_DEFAULT_SECRET", "false").lower() != "true":
    raise RuntimeError(
        "ADMIN_PASS is the known insecure default. Set a strong ADMIN_PASS env var, "
        "or set ALLOW_DEFAULT_SECRET=true for local development only."
    )
ALLOW_REGISTRATION = os.environ.get("ALLOW_REGISTRATION", "false").lower() == "true"


# ── DB ────────────────────────────────────────────────────────
def get_db():
    """Get a Postgres connection from the pool."""
    return database_module.get_db_sync()


# ── Event logging ────────────────────────────────────────────────
def log_event(user_id: int, username: str, event_type: str, detail: str = None, ip: str = None):
    """Fire-and-forget event logger. Never raises."""
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO events (user_id, username, event_type, detail, ip) VALUES (%s,%s,%s,%s,%s)",
            (user_id, username, event_type, detail, ip)
        )
        conn.commit()
        conn.close()
    except Exception:
        pass


# ── Auth ──────────────────────────────────────────────────────
def verify_password(plain: str, hashed: str) -> bool:
    return _bcrypt.checkpw(plain.encode(), hashed.encode())

def hash_password(plain: str) -> str:
    return _bcrypt.hashpw(plain.encode(), _bcrypt.gensalt()).decode()

def validate_password(pw: str):
    """Raise HTTPException if password doesn't meet complexity requirements."""
    errors = []
    if len(pw) < 10:
        errors.append("at least 10 characters")
    if not re.search(r"[A-Z]", pw):
        errors.append("one uppercase letter")
    if not re.search(r"[a-z]", pw):
        errors.append("one lowercase letter")
    if not re.search(r"\d", pw):
        errors.append("one number")
    if not re.search(r"[^A-Za-z0-9]", pw):
        errors.append("one special character")
    if errors:
        raise HTTPException(status_code=400, detail="Password requires: " + ", ".join(errors))


# ── AI key encryption (Fernet, derived from SECRET_KEY) ───────
import base64 as _b64, hashlib as _hashlib
from cryptography.fernet import Fernet, InvalidToken

def _fernet():
    # Derive a stable 32-byte urlsafe key from SECRET_KEY
    digest = _hashlib.sha256(SECRET_KEY.encode()).digest()
    return Fernet(_b64.urlsafe_b64encode(digest))

def encrypt_secret(plain: str) -> str:
    return _fernet().encrypt(plain.encode()).decode()

def decrypt_secret(token: str) -> str:
    try:
        return _fernet().decrypt(token.encode()).decode()
    except InvalidToken:
        return ""

def create_token(data: dict):
    payload = data.copy()
    payload["exp"] = datetime.utcnow() + timedelta(days=TOKEN_EXPIRE_DAYS)
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def get_user(conn, username: str):
    return conn.execute("SELECT * FROM users WHERE username=%s", (username.strip().lower(),)).fetchone()

async def current_user(token: str = Depends(oauth2)):
    err = HTTPException(status_code=401, detail="Invalid or expired token",
                        headers={"WWW-Authenticate": "Bearer"})
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if not username:
            raise err
    except JWTError:
        raise err
    conn = get_db()
    user = get_user(conn, username)
    conn.close()
    if not user:
        raise err
    return dict(user)

async def admin_only(user=Depends(current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

def data_user_id(user: dict) -> int:
    if user["role"] == "demo":
        conn = get_db()
        admin = conn.execute("SELECT id FROM users WHERE role='admin' LIMIT 1").fetchone()
        conn.close()
        return admin["id"] if admin else 1
    return user["id"]


# ── Models ────────────────────────────────────────────────────
class SetData(BaseModel):
    reps: float
    rawLoad: float
    trueLbs: float
    vol: float
    time: float = 1.5          # per-set time in minutes

class ExerciseData(BaseModel):
    name: str
    sets: list[SetData]
    totalVol: float
    density: float
    tool: Optional[str] = "Bar"
    mult: Optional[float] = 2.0
    isBW: Optional[bool] = False
    time: Optional[float] = None  # legacy field — per-set time is now in SetData.time

class ExerciseIn(BaseModel):
    name: str
    alias: Optional[str] = ''
    tool: str = 'Bar'
    mult: float = 2.0
    muscles: list[str] = []
    day: Optional[str] = ''
    load_hint: Optional[str] = ''
    is_bw: bool = False
    sort_order: int = 0
    rep_trigger_override: Optional[int] = None

class SessionIn(BaseModel):
    date: str
    bw: float
    rd: float
    total_density: float
    exercises: list[ExerciseData]
    notes: Optional[str] = ''
    sleep_hours: Optional[float] = None
    deep_sleep_pct: Optional[int] = None

class UserCreate(BaseModel):
    username: str
    password: str
    role: str = "guest"


class ProfileIn(BaseModel):
    first_name: Optional[str] = None
    last_name:  Optional[str] = None
    dob:        Optional[str] = None        # YYYY-MM-DD
    gender:     Optional[str] = None
    week_start: Optional[str] = None        # Monday-Sunday
    height_in:      Optional[float] = None      # inches
    target_bw:      Optional[float] = None      # lbs
    activity_level: Optional[str]   = None      # TDEE multiplier string
    onboarded:          Optional[str] = None
    last_seen_version:  Optional[str] = None
    external_api_key:   Optional[str] = None
    rep_trigger:        Optional[int] = None    # cumulative rep threshold (default 50)
    set_time:           Optional[float] = None  # minutes per set (default 1.5)
    ollama_base_url:    Optional[str] = None    # e.g. http://localhost:11434

class ProtocolIn(BaseModel):
    name:       str
    dose:       Optional[str] = None
    frequency:  Optional[str] = None
    notes:      Optional[str] = None
    start_date: Optional[str] = None
    end_date:   Optional[str] = None
    track:      Optional[bool] = False

class ProtocolTrack(BaseModel):
    track: Optional[bool] = None
    end_date: Optional[str] = None

class InventoryItemIn(BaseModel):
    name:         str
    item_type:    Optional[str] = "vial"     # vial | capsule | powder | nasal | other
    protocol_id:  Optional[int] = None
    total_amount: Optional[float] = None      # e.g. 80 (mg), 60 (caps), 500 (g)
    unit:         Optional[str] = "mg"        # mg | iu | caps | g | sprays
    bac_water_ml: Optional[float] = None      # vials only; enables units-math
    remaining:    Optional[float] = None      # defaults to total_amount
    per_dose:     Optional[float] = None      # default deduction per dose event
    opened_date:  Optional[str] = None
    status:       Optional[str] = "sealed"    # sealed | active | empty
    vendor:       Optional[str] = None
    cost:         Optional[float] = None
    notes:        Optional[str] = None

class DoseEventIn(BaseModel):
    protocol_id: Optional[int] = None
    item_id:     Optional[int] = None         # if set, decrements item remaining
    amount:      Optional[float] = None       # falls back to item per_dose
    unit:        Optional[str] = None         # falls back to item unit
    taken_at:    Optional[str] = None         # ISO timestamp; defaults to now
    site:        Optional[str] = None         # injection site, optional
    notes:       Optional[str] = None

class PhaseIn(BaseModel):
    phase_type: str
    start_date: str
    end_date:   Optional[str] = None
    notes:      Optional[str] = None
    label:      Optional[str] = None

class AISettingsIn(BaseModel):
    ai_key: Optional[str] = None      # plaintext key from user; None = don't change
    ai_model: Optional[str] = None
    ollama_base_url: Optional[str] = None  # e.g., http://localhost:11434

class MFAVerify(BaseModel):
    code: str

class PasswordSelfChange(BaseModel):
    current_password: str
    new_password: str

class PasswordChange(BaseModel):
    password: str

class UserRegister(BaseModel):
    username: str
    password: str

class InsightsQuery(BaseModel):
    question: str


# ── Auth endpoints ────────────────────────────────────────────
@app.get("/health")
def health():
    return {"status": "ok"}

@app.get("/auth/registration-status")
def registration_status():
    """Public — lets the frontend know whether to show the sign-up link."""
    return {"open": ALLOW_REGISTRATION}

@app.post("/auth/register")
def register(body: UserRegister, request: Request):
    """Self-service account creation. Only active when ALLOW_REGISTRATION=true."""
    if not ALLOW_REGISTRATION:
        raise HTTPException(status_code=403, detail="Registration is not open")

    username = body.username.strip().lower()
    if not username:
        raise HTTPException(status_code=400, detail="Email address is required")
    if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', username):
        raise HTTPException(status_code=400, detail="Please enter a valid email address")

    validate_password(body.password)

    conn = get_db()
    if conn.execute("SELECT id FROM users WHERE username=%s", (username,)).fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Username already taken")

    conn.execute(
        "INSERT INTO users (username, hashed_pw, role, status) VALUES (%s,%s,%s,%s)",
        (username, hash_password(body.password), "guest", "pending")
    )
    conn.commit()
    user = get_user(conn, username)
    conn.close()

    log_event(user["id"], user["username"], "register",
              "pending approval", request.client.host if request.client else None)
    return {"pending": True,
            "message": "Account created. You'll be able to sign in once an admin approves your request."}

@app.post("/auth/login")
def login(request: Request, form: OAuth2PasswordRequestForm = Depends()):
    conn = get_db()
    user = get_user(conn, form.username.strip().lower())
    if not user or not verify_password(form.password, user["hashed_pw"]):
        conn.close()
        raise HTTPException(status_code=401, detail="Incorrect username or password")

    status = user["status"] if "status" in user.keys() else "active"
    if status == "pending":
        conn.close()
        raise HTTPException(status_code=403, detail="Your account is pending admin approval.")
    if status == "suspended":
        conn.close()
        raise HTTPException(status_code=403, detail="Your account has been suspended. Contact an admin.")

    # Check if MFA is enabled for this user
    settings = conn.execute(
        "SELECT totp_enabled FROM user_settings WHERE user_id=%s", (user["id"],)
    ).fetchone()
    conn.close()

    mfa_enabled = settings and settings["totp_enabled"] == "1"

    if mfa_enabled:
        # Issue a short-lived MFA pending token (5 min), no app access
        pending_token = create_token({
            "sub": user["username"],
            "role": user["role"],
            "mfa_pending": True,
            "exp": datetime.utcnow() + timedelta(minutes=5),
        })
        log_event(user["id"], user["username"], "login_mfa_pending",
                  f"role={user['role']}", request.client.host if request.client else None)
        return {"mfa_required": True, "pending_token": pending_token}

    token = create_token({"sub": user["username"], "role": user["role"]})
    log_event(user["id"], user["username"], "login",
              f"role={user['role']}", request.client.host if request.client else None)
    return {"access_token": token, "token_type": "bearer",
            "role": user["role"], "username": user["username"]}


@app.post("/auth/mfa/verify")
def mfa_verify(body: MFAVerify, request: Request, token: str = Depends(oauth2)):
    """Second step: verify TOTP code using the pending token, return full JWT."""
    err = HTTPException(status_code=401, detail="Invalid or expired MFA session")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError:
        raise err
    if not payload.get("mfa_pending"):
        raise HTTPException(status_code=400, detail="Token is not an MFA pending token")

    username = payload.get("sub")
    conn = get_db()
    user = get_user(conn, username)
    if not user:
        conn.close()
        raise err

    settings = conn.execute(
        "SELECT totp_secret FROM user_settings WHERE user_id=%s", (user["id"],)
    ).fetchone()
    conn.close()

    if not settings or not settings["totp_secret"]:
        raise HTTPException(status_code=400, detail="MFA not configured")

    secret = decrypt_secret(settings["totp_secret"])
    totp = pyotp.TOTP(secret)
    if not totp.verify(body.code.strip(), valid_window=1):
        raise HTTPException(status_code=401, detail="Invalid TOTP code")

    full_token = create_token({"sub": user["username"], "role": user["role"]})
    log_event(user["id"], user["username"], "login_mfa_ok",
              f"role={user['role']}", request.client.host if request.client else None)
    return {"access_token": full_token, "token_type": "bearer",
            "role": user["role"], "username": user["username"]}


@app.get("/auth/mfa/setup")
def mfa_setup(user=Depends(current_user)):
    """Generate a new TOTP secret and return QR code PNG as base64. Does NOT enable MFA yet."""
    secret = pyotp.random_base32()
    totp = pyotp.TOTP(secret)
    label = f"Agon:{user['username']}"
    uri = totp.provisioning_uri(name=label, issuer_name="Agon")

    img = qrcode.make(uri)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    # Store the (unconfirmed) secret encrypted — will be activated on /enable
    enc = encrypt_secret(secret)
    conn = get_db()
    existing = conn.execute("SELECT user_id FROM user_settings WHERE user_id=%s", (user["id"],)).fetchone()
    if existing:
        conn.execute("UPDATE user_settings SET totp_secret=%s, totp_enabled='0' WHERE user_id=%s",
                     (enc, user["id"]))
    else:
        conn.execute("INSERT INTO user_settings (user_id, totp_secret, totp_enabled) VALUES (%s,%s,'0')",
                     (user["id"], enc))
    conn.commit()
    conn.close()
    return {"qr": qr_b64, "secret": secret, "uri": uri}


@app.post("/auth/mfa/enable")
def mfa_enable(body: MFAVerify, user=Depends(current_user)):
    """Confirm TOTP code is working, then flip totp_enabled = 1."""
    conn = get_db()
    settings = conn.execute(
        "SELECT totp_secret FROM user_settings WHERE user_id=%s", (user["id"],)
    ).fetchone()
    conn.close()

    if not settings or not settings["totp_secret"]:
        raise HTTPException(status_code=400, detail="Run /auth/mfa/setup first")

    secret = decrypt_secret(settings["totp_secret"])
    totp = pyotp.TOTP(secret)
    if not totp.verify(body.code.strip(), valid_window=1):
        raise HTTPException(status_code=401, detail="Invalid TOTP code — try again")

    conn = get_db()
    conn.execute("UPDATE user_settings SET totp_enabled='1' WHERE user_id=%s", (user["id"],))
    conn.commit()
    conn.close()
    log_event(user["id"], user["username"], "mfa_enabled", "TOTP activated")
    return {"status": "mfa_enabled"}


@app.post("/auth/mfa/disable")
def mfa_disable(user=Depends(current_user)):
    """Any user: disable their own MFA."""
    conn = get_db()
    conn.execute(
        "UPDATE user_settings SET totp_enabled='0', totp_secret=NULL WHERE user_id=%s",
        (user["id"],)
    )
    conn.commit()
    conn.close()
    log_event(user["id"], user["username"], "mfa_disabled", "TOTP removed")
    return {"status": "mfa_disabled"}


@app.post("/admin/users/{username}/mfa/disable")
def admin_disable_user_mfa(username: str, user=Depends(admin_only)):
    """Admin: disable MFA for any user by username."""
    conn = get_db()
    target = conn.execute("SELECT id FROM users WHERE username=%s", (username,)).fetchone()
    if not target:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")
    conn.execute(
        "UPDATE user_settings SET totp_enabled='0', totp_secret=NULL WHERE user_id=%s",
        (target["id"],)
    )
    conn.commit()
    conn.close()
    log_event(user["id"], user["username"], "mfa_disabled", f"admin disabled MFA for {username}")
    return {"status": "mfa_disabled", "username": username}


@app.get("/auth/me")
def me(user=Depends(current_user)):
    conn = get_db()
    s = conn.execute(
        "SELECT totp_enabled FROM user_settings WHERE user_id=%s", (user["id"],)
    ).fetchone()
    conn.close()
    mfa_enabled = bool(s and s["totp_enabled"] == "1")
    return {"username": user["username"], "role": user["role"], "id": user["id"],
            "mfa_enabled": mfa_enabled}


@app.put("/auth/password")
def change_own_password(body: PasswordSelfChange, user=Depends(current_user)):
    """Any authenticated user can change their own password by providing their current one."""
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot change their password")
    conn = get_db()
    stored = conn.execute("SELECT hashed_pw FROM users WHERE id=%s", (user["id"],)).fetchone()
    conn.close()
    if not stored or not verify_password(body.current_password, stored["hashed_pw"]):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    validate_password(body.new_password)
    conn = get_db()
    conn.execute("UPDATE users SET hashed_pw=%s WHERE id=%s",
                 (hash_password(body.new_password), user["id"]))
    conn.commit()
    conn.close()
    log_event(user["id"], user["username"], "password_changed", "self-service")
    return {"status": "updated"}


# ── Admin ─────────────────────────────────────────────────────
@app.get("/admin/users")
def list_users(user=Depends(admin_only)):
    conn = get_db()
    rows = conn.execute("SELECT id, username, role, status, created_at FROM users ORDER BY id").fetchall()
    result = []
    for r in rows:
        s = conn.execute(
            "SELECT totp_enabled FROM user_settings WHERE user_id=%s", (r["id"],)
        ).fetchone()
        result.append({**dict(r), "mfa_enabled": bool(s and s["totp_enabled"] == "1")})
    conn.close()
    return result


@app.post("/admin/users/{username}/approve")
def approve_user(username: str, user=Depends(admin_only)):
    conn = get_db()
    target = conn.execute("SELECT id, status FROM users WHERE username=%s", (username,)).fetchone()
    if not target:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")
    conn.execute("UPDATE users SET status='active' WHERE username=%s", (username,))
    conn.commit()
    conn.close()
    log_event(user["id"], user["username"], "user_approved", f"approved {username}")
    return {"status": "approved", "username": username}


@app.post("/admin/users/{username}/reject")
def reject_user(username: str, user=Depends(admin_only)):
    conn = get_db()
    target = conn.execute("SELECT id FROM users WHERE username=%s", (username,)).fetchone()
    if not target:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")
    conn.execute("DELETE FROM users WHERE username=%s", (username,))
    conn.commit()
    conn.close()
    log_event(user["id"], user["username"], "user_rejected", f"rejected and deleted {username}")
    return {"status": "rejected", "username": username}

@app.post("/admin/users")
def create_user(body: UserCreate, user=Depends(admin_only)):
    validate_password(body.password)
    conn = get_db()
    body.username = body.username.strip().lower()
    if conn.execute("SELECT id FROM users WHERE username=%s", (body.username,)).fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Username already exists")
    conn.execute("INSERT INTO users (username, hashed_pw, role) VALUES (%s,%s,%s)",
                 (body.username, hash_password(body.password), body.role))
    conn.commit()
    conn.close()
    return {"status": "created", "username": body.username}

@app.delete("/admin/users/{username}")
def delete_user(username: str, user=Depends(admin_only)):
    if username == user["username"]:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    conn = get_db()
    target = conn.execute("SELECT role FROM users WHERE username=%s", (username,)).fetchone()
    if not target:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")
    if target["role"] == "admin":
        conn.close()
        raise HTTPException(status_code=400, detail="Cannot delete an admin account")
    conn.execute("DELETE FROM users WHERE username=%s", (username,))
    conn.commit()
    conn.close()
    return {"status": "deleted", "username": username}

@app.put("/admin/users/{username}/password")
def change_password(username: str, body: PasswordChange, user=Depends(admin_only)):
    validate_password(body.password)
    conn = get_db()
    target = conn.execute("SELECT id FROM users WHERE username=%s", (username,)).fetchone()
    if not target:
        conn.close()
        raise HTTPException(status_code=404, detail="User not found")
    conn.execute("UPDATE users SET hashed_pw=%s WHERE username=%s",
                 (hash_password(body.password), username))
    conn.commit()
    conn.close()
    return {"status": "updated"}


# ── Sessions ──────────────────────────────────────────────────
@app.get("/sessions")
def get_sessions(user=Depends(current_user)):
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM sessions WHERE user_id=%s ORDER BY date ASC", (uid,)
    ).fetchall()
    conn.close()
    return [{"id": r["id"], "date": r["date"], "bw": r["bw"], "rd": r["rd"],
             "total_density": r["total_density"], "notes": r["notes"] or "",
             "sleep_hours": r.get("sleep_hours"), "deep_sleep_pct": r.get("deep_sleep_pct"),
             "exercises": r["exercises"] if isinstance(r["exercises"], list) else json.loads(r["exercises"])} for r in rows]

@app.post("/sessions")
def save_session(session: SessionIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo account is read-only")
    uid = user["id"]

    # Recalculate RD server-side using correct formula:
    # RD = total_session_vol / (total_session_time * bodyweight)
    exercises_out = []
    total_vol  = 0.0
    total_time = 0.0
    for ex in session.exercises:
        ex_vol  = sum(s.vol  for s in ex.sets)
        ex_time = sum(s.time for s in ex.sets if s.time > 0)
        total_vol  += ex_vol
        total_time += ex_time
        exercises_out.append({
            **ex.dict(),
            "totalVol": round(ex_vol, 1),
            "density":  round(ex_vol / ex_time, 2) if ex_time > 0 else 0,
        })

    total_density = round(total_vol / total_time, 2) if total_time > 0 else 0
    rd = round(total_vol / (total_time * session.bw), 2) if (total_time > 0 and session.bw) else 0

    conn = get_db()
    try:
        # Try UPDATE first
        cursor = conn.execute(
            """UPDATE sessions SET bw=%s, rd=%s, total_density=%s, exercises=%s, notes=%s, sleep_hours=%s, deep_sleep_pct=%s
               WHERE user_id=%s AND date=%s""",
            (session.bw, rd, round(total_density, 2), json.dumps(exercises_out), 
             session.notes or "", session.sleep_hours, session.deep_sleep_pct, uid, session.date)
        )
        
        # If no rows were updated, INSERT a new one
        if cursor.rowcount == 0:
            conn.execute(
                """INSERT INTO sessions (user_id, date, bw, rd, total_density, exercises, notes, sleep_hours, deep_sleep_pct)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (uid, session.date, session.bw, rd, round(total_density, 2), 
                 json.dumps(exercises_out), session.notes or "", session.sleep_hours, session.deep_sleep_pct)
            )
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))
    conn.close()
    log_event(user["id"], user["username"], "session_saved", f"date={session.date} rd={rd:.2f}")
    return {"status": "saved", "date": session.date, "rd": rd}

@app.delete("/sessions/{date}")
def delete_session(date: str, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo account is read-only")
    conn = get_db()
    conn.execute("DELETE FROM sessions WHERE user_id=%s AND date=%s", (user["id"], date))
    conn.commit()
    conn.close()
    return {"status": "deleted", "date": date}


# ── Trend ─────────────────────────────────────────────────────
@app.get("/trend")
def get_trend(user=Depends(current_user)):
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT date, bw, rd FROM sessions WHERE user_id=%s ORDER BY date ASC", (uid,)
    ).fetchall()
    settings_row = conn.execute("SELECT week_start FROM user_settings WHERE user_id=%s", (uid,)).fetchone()
    conn.close()
    week_start_pref = (settings_row["week_start"] if settings_row and settings_row["week_start"] else "Saturday")

    # weekday(): Mon=0, Tue=1, Wed=2, Thu=3, Fri=4, Sat=5, Sun=6
    # Week ends on the day BEFORE week_start
    DAY_MAP = {"Monday":0,"Tuesday":1,"Wednesday":2,"Thursday":3,"Friday":4,"Saturday":5,"Sunday":6}
    start_wd = DAY_MAP.get(week_start_pref, 5)  # default Saturday=5
    end_wd = (start_wd - 1) % 7  # week-end day

    by_week: dict = {}
    for r in rows:
        try:
            # Handle both Postgres (datetime.date) and SQLite (string) date formats
            if isinstance(r["date"], str):
                d = datetime.strptime(r["date"], "%Y-%m-%d")
            else:
                # r["date"] is already a datetime.date object from Postgres
                d = datetime.combine(r["date"], datetime.min.time())
        except (ValueError, TypeError):
            continue
        delta = (end_wd - d.weekday()) % 7
        week_end = d + timedelta(days=delta)
        key = f"{week_end.month}/{week_end.day}/{week_end.year}"
        if key not in by_week:
            by_week[key] = []
        by_week[key].append({"rd": r["rd"], "wt": r["bw"]})

    result = []
    for week, entries in by_week.items():
        result.append({
            "week": week,
            "rd":   round(sum(e["rd"] for e in entries) / len(entries), 2),
            "wt":   round(sum(e["wt"] for e in entries) / len(entries), 1),
        })
    result.sort(key=lambda w: datetime.strptime(w["week"], "%m/%d/%Y"))
    return result


# ── Progress ──────────────────────────────────────────────────
@app.get("/progress/{exercise_name}")
def get_progress(exercise_name: str, user=Depends(current_user)):
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT date, bw, rd, exercises FROM sessions WHERE user_id=%s ORDER BY date ASC", (uid,)
    ).fetchall()
    conn.close()

    result = []
    for r in rows:
        exercises = r["exercises"] if isinstance(r["exercises"], list) else json.loads(r["exercises"])
        match = next((e for e in exercises
                      if e.get("name","").lower() == exercise_name.lower()), None)
        if not match:
            continue
        sets = match.get("sets", [])
        if not sets:
            continue
        top_set  = max(sets, key=lambda s: s.get("trueLbs", 0))
        total_vol = sum(s.get("vol", 0) for s in sets)
        result.append({
            "date":     r["date"],
            "bw":       r["bw"],
            "rd":       r["rd"],
            "topLoad":  round(top_set.get("trueLbs", 0), 1),
            "topReps":  int(top_set.get("reps", 0)),
            "totalVol": round(total_vol, 1),
            "density":  round(match.get("density", 0), 2),
            "sets":     len(sets),
        })
    return result


@app.get("/exercises/bank")
def get_exercise_bank(user=Depends(current_user)):
    conn = get_db()
    rows = conn.execute("SELECT * FROM exercises ORDER BY name ASC").fetchall()
    conn.close()
    return [{"id":r["id"],"name":r["name"],"alias":r["alias"],"tool":r["tool"],
             "mult":r["mult"],"muscles":r["muscles"] if isinstance(r["muscles"], list) else json.loads(r["muscles"]),"day":r["day"],
             "loadHint":r["load_hint"],"isBW":bool(r["is_bw"]),"sortOrder":r["sort_order"],
             "createdBy":r["created_by"],
             "repTriggerOverride": r["rep_trigger_override"] if r["rep_trigger_override"] else None
             } for r in rows]

@app.post("/exercises/bank")
def add_exercise(body: ExerciseIn, user=Depends(current_user)):
    if user["role"] not in ("admin", "guest"):
        raise HTTPException(status_code=403, detail="Read-only account")
    if not body.muscles or len(body.muscles) == 0:
        raise HTTPException(status_code=400, detail="At least one muscle group is required")
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO exercises (name, alias, tool, mult, muscles, day, load_hint, is_bw, sort_order, created_by, rep_trigger_override)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT(name) DO UPDATE SET
                alias=excluded.alias, tool=excluded.tool, mult=excluded.mult,
                muscles=excluded.muscles, day=excluded.day, load_hint=excluded.load_hint,
                is_bw=excluded.is_bw, sort_order=excluded.sort_order,
                rep_trigger_override=excluded.rep_trigger_override
        """, (body.name, body.alias, body.tool, body.mult, json.dumps(body.muscles),
              body.day, body.load_hint, body.is_bw, body.sort_order, user["id"],
              body.rep_trigger_override))
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))
    conn.close()
    return {"status": "upserted", "name": body.name}

@app.put("/exercises/bank/{ex_name}")
def update_exercise(ex_name: str, body: ExerciseIn, user=Depends(current_user)):
    if user["role"] not in ("admin", "guest"):
        raise HTTPException(status_code=403, detail="Read-only account")
    conn = get_db()
    existing = conn.execute("SELECT created_by FROM exercises WHERE name=%s", (ex_name,)).fetchone()
    if existing:
        # Non-admin can only edit their own exercises
        if user["role"] != "admin" and existing["created_by"] != user["id"]:
            conn.close()
            raise HTTPException(status_code=403, detail="You can only edit exercises you created")
        result = conn.execute(
            "UPDATE exercises SET alias=%s, tool=%s, mult=%s, muscles=%s, day=%s, load_hint=%s, is_bw=%s, sort_order=%s, rep_trigger_override=%s WHERE name=%s",
            (body.alias, body.tool, body.mult, json.dumps(body.muscles),
             body.day, body.load_hint, body.is_bw, body.sort_order,
             body.rep_trigger_override, ex_name))
    else:
        # New exercise — insert with ownership
        conn.execute(
            "INSERT INTO exercises (name, alias, tool, mult, muscles, day, load_hint, is_bw, sort_order, created_by, rep_trigger_override) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (ex_name, body.alias, body.tool, body.mult, json.dumps(body.muscles),
             body.day, body.load_hint, body.is_bw, body.sort_order, user["id"],
             body.rep_trigger_override))
    conn.commit()
    conn.close()
    return {"status": "upserted"}

@app.delete("/exercises/bank/{ex_name}")
def delete_exercise(ex_name: str, user=Depends(current_user)):
    if user["role"] not in ("admin", "guest"):
        raise HTTPException(status_code=403, detail="Read-only account")
    conn = get_db()
    existing = conn.execute("SELECT created_by FROM exercises WHERE name=%s", (ex_name,)).fetchone()
    if not existing:
        conn.close()
        return {"status": "not found"}
    if user["role"] != "admin" and existing["created_by"] != user["id"]:
        conn.close()
        raise HTTPException(status_code=403, detail="You can only delete exercises you created")
    conn.execute("DELETE FROM exercises WHERE name=%s", (ex_name,))
    conn.commit()
    conn.close()
    return {"status": "deleted"}


@app.post("/admin/reset-sequences")
def reset_sequences(user=Depends(admin_only)):
    """Reset all SERIAL sequences to MAX(id)+1 to fix duplicate-key errors after data migrations."""
    conn = get_db()
    tables = ["users", "sessions", "exercises", "protocols", "phases",
              "insights_messages", "events"]
    results = {}
    for t in tables:
        try:
            row = conn.execute(f"SELECT setval('{t}_id_seq', (SELECT COALESCE(MAX(id), 0) + 1 FROM {t}))").fetchone()
            results[t] = row[0] if row else None
        except Exception as e:
            results[t] = f"error: {e}"
    conn.commit()
    conn.close()
    return {"status": "ok", "sequences": results}


@app.get("/exercises/{ex_name}/set-curve")
def get_set_curve(ex_name: str, user=Depends(current_user)):
    """Per-set rep data for fixed-load accessory blocks (2–9 sets, all same load ±0.5 lbs)."""
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT date, exercises FROM sessions WHERE user_id=%s ORDER BY date ASC", (uid,)
    ).fetchall()
    conn.close()

    curve_sessions = []
    load_history   = []

    for row in rows:
        exs = row["exercises"] if isinstance(row["exercises"], list) else json.loads(row["exercises"])
        for ex in exs:
            if ex.get("name", "").lower() != ex_name.lower():
                continue
            sets = ex.get("sets", [])
            if len(sets) < 2 or len(sets) > 10:
                continue
            loads = [s.get("trueLbs") or s.get("rawLoad") or 0 for s in sets]
            if not all(abs(l - loads[0]) < 0.5 for l in loads):
                continue
            reps_per_set = [s.get("reps", 0) for s in sets]
            total_reps   = sum(reps_per_set)
            block_load   = round(loads[0], 1)
            curve_sessions.append({
                "date":         row["date"],
                "load":         block_load,
                "reps_per_set": reps_per_set,
                "total_reps":   total_reps,
                "set_count":    len(sets),
                "floor_rep":    min(reps_per_set),
                "opening_rep":  reps_per_set[0] if reps_per_set else 0,
            })
            load_history.append({"date": row["date"], "load": block_load})

    return {"exercise": ex_name, "sessions": curve_sessions, "load_history": load_history}


@app.get("/exercises/logged")
def get_logged_exercises(user=Depends(current_user)):
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT exercises FROM sessions WHERE user_id=%s", (uid,)
    ).fetchall()
    conn.close()
    names = set()
    for r in rows:
        exercises = r["exercises"] if isinstance(r["exercises"], list) else json.loads(r["exercises"])
        for ex in exercises:
            if ex.get("name"):
                names.add(ex["name"])
    return sorted(names)

# ── Profile endpoints ────────────────────────────────────────
@app.get("/profile")
def get_profile(user=Depends(current_user)):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT first_name, last_name, dob, gender, week_start, height_in, target_bw, activity_level, onboarded, last_seen_version, ollama_base_url, set_time FROM user_settings WHERE user_id=%s",
            (user["id"],)
        ).fetchone()
    except Exception:
        # Columns may not exist yet — fall back to partial query
        try:
            row = conn.execute(
                "SELECT first_name, last_name, dob, gender, week_start FROM user_settings WHERE user_id=%s",
                (user["id"],)
            ).fetchone()
        except Exception:
            row = None
    conn.close()
    if not row:
        return {"first_name":"","last_name":"","dob":"","gender":"","week_start":"Saturday","height_in":None,"target_bw":None,"ollama_base_url":""}
    return {
        "first_name": row["first_name"] or "",
        "last_name":  row["last_name"]  or "",
        "dob":        row["dob"]        or "",
        "gender":     row["gender"]     or "",
        "week_start":       row["week_start"] or "Saturday",
        "height_in":        row["height_in"]        if "height_in"        in row.keys() else None,
        "target_bw":        row["target_bw"]        if "target_bw"        in row.keys() else None,
        "activity_level":   row["activity_level"]   if "activity_level"   in row.keys() else "1.55",
        "onboarded":        row["onboarded"]        if "onboarded"        in row.keys() else "0",
        "last_seen_version": row["last_seen_version"] if "last_seen_version" in row.keys() else "0.0.0",
        "rep_trigger":      int(row["rep_trigger"]) if "rep_trigger" in row.keys() and row["rep_trigger"] else 50,
        "set_time":         float(row["set_time"]) if "set_time" in row.keys() and row["set_time"] else 1.5,
        "ollama_base_url":  row["ollama_base_url"]  if "ollama_base_url"  in row.keys() else "",
    }

@app.post("/profile")
def save_profile(body: ProfileIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot save profile data")
    # ollama_base_url is a server-side request target (SSRF sink). Only admins may
    # set it; for everyone else, ignore the field rather than failing the whole save.
    if body.ollama_base_url and body.ollama_base_url.strip() and user["role"] != "admin":
        body.ollama_base_url = None
    conn = get_db()
    existing = conn.execute("SELECT user_id FROM user_settings WHERE user_id=%s", (user["id"],)).fetchone()
    
    # Build dynamic UPDATE to only update provided fields
    if existing:
        updates = []
        params = []
        field_map = {
            'first_name': body.first_name,
            'last_name': body.last_name,
            'dob': body.dob,
            'gender': body.gender,
            'week_start': body.week_start,
            'height_in': body.height_in,
            'target_bw': body.target_bw,
            'activity_level': body.activity_level,
            'onboarded': body.onboarded,
            'last_seen_version': body.last_seen_version,
            'rep_trigger': body.rep_trigger,
            'set_time': body.set_time,
            'ollama_base_url': body.ollama_base_url,
        }
        for field, value in field_map.items():
            if value is not None:
                updates.append(f"{field}=%s")
                params.append(value)
        
        if updates:
            updates.append("updated_at=CURRENT_TIMESTAMP")
            sql = f"UPDATE user_settings SET {', '.join(updates)} WHERE user_id=%s"
            params.append(user["id"])
            conn.execute(sql, params)
    else:
        conn.execute("""
            INSERT INTO user_settings (user_id, first_name, last_name, dob, gender, week_start, height_in, target_bw, activity_level, onboarded, last_seen_version, rep_trigger)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (user["id"], body.first_name, body.last_name, body.dob, body.gender,
                body.week_start or "Saturday", body.height_in, body.target_bw,
                body.activity_level or "1.55", body.onboarded or "0",
                body.last_seen_version or "0.0.0", body.rep_trigger or 50))
    conn.commit()
    conn.close()
    return {"status": "saved"}


# ── AI Settings endpoints ─────────────────────────────────────
@app.get("/ai/settings")
def get_ai_settings(user=Depends(current_user)):
    """Return whether a key is set + masked preview + model + ollama URL."""
    conn = get_db()
    row = conn.execute("SELECT ai_key_enc, ai_model, ollama_base_url FROM user_settings WHERE user_id=%s",
                       (user["id"],)).fetchone()
    conn.close()
    if not row or not row["ai_key_enc"]:
        return {
            "has_key": False,
            "masked": "",
            "model": (row["ai_model"] if row else "gemini-2.5-flash"),
            "ollama_base_url": (row["ollama_base_url"] if row else "")
        }
    plain = decrypt_secret(row["ai_key_enc"])
    masked = ("•" * max(0, len(plain) - 4) + plain[-4:]) if plain else ""
    return {
        "has_key": bool(plain),
        "masked": masked,
        "model": row["ai_model"] or "gemini-2.5-flash",
        "ollama_base_url": row["ollama_base_url"] or ""
    }

@app.post("/ai/settings")
def save_ai_settings(body: AISettingsIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot store API keys")
    # The Ollama base URL is a server-side request target (SSRF sink): the backend
    # POSTs to whatever host it names. Only admins may set it. Clearing it is fine.
    if body.ollama_base_url and body.ollama_base_url.strip() and user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only an admin can set the Ollama base URL")
    conn = get_db()
    existing = conn.execute("SELECT user_id FROM user_settings WHERE user_id=%s",
                            (user["id"],)).fetchone()
    # Build the update — always write fields atomically
    key_enc = None
    if body.ai_key is not None and body.ai_key.strip():
        key_enc = encrypt_secret(body.ai_key.strip())
    if existing:
        cur = conn.execute("SELECT ai_key_enc, ai_model, ollama_base_url FROM user_settings WHERE user_id=%s",
                           (user["id"],)).fetchone()
        final_key        = key_enc if key_enc is not None else cur["ai_key_enc"]
        final_model      = body.ai_model if body.ai_model else (cur["ai_model"] or "gemini-2.5-flash")
        final_ollama_url = body.ollama_base_url if body.ollama_base_url is not None else cur["ollama_base_url"]
        conn.execute(
            "UPDATE user_settings SET ai_key_enc=%s, ai_model=%s, ollama_base_url=%s, updated_at=CURRENT_TIMESTAMP WHERE user_id=%s",
            (final_key, final_model, final_ollama_url, user["id"])
        )
    else:
        conn.execute("INSERT INTO user_settings (user_id, ai_key_enc, ai_model, ollama_base_url) VALUES (%s,%s,%s,%s)",
                     (user["id"], key_enc, body.ai_model or "gemini-2.5-flash", body.ollama_base_url or ""))
    conn.commit()
    conn.close()
    return {"status": "saved"}

@app.delete("/ai/settings")
def delete_ai_settings(user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts have no stored key")
    conn = get_db()
    conn.execute("UPDATE user_settings SET ai_key_enc=NULL, updated_at=CURRENT_TIMESTAMP WHERE user_id=%s",
                 (user["id"],))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

# ── Insights: build a compact data context for the LLM ────────
def build_training_context(uid: int) -> str:
    """Produce a compact text summary of the user's training data for the LLM."""
    conn = get_db()
    rows = conn.execute(
        "SELECT date, bw, rd, total_density, exercises, notes, sleep_hours, deep_sleep_pct FROM sessions WHERE user_id=%s ORDER BY date",
        (uid,)
    ).fetchall()
    if not rows:
        conn.close()
        return "No training sessions logged yet."

    lines = []
    # Load profile for richer AI context
    prof = conn.execute(
        "SELECT first_name, last_name, dob, gender, week_start, height_in, target_bw FROM user_settings WHERE user_id=%s",
        (uid,)
    ).fetchone()

    def prof_get(key, default=None):
        try:
            return prof[key]
        except (IndexError, TypeError):
            return default

    if prof:
        name_str = " ".join(filter(None, [prof_get("first_name"), prof_get("last_name")])) or "User"
        age_str = ""
        if prof_get("dob"):
            try:
                dob = datetime.strptime(prof_get("dob"), "%Y-%m-%d").date()
                age_str = f", age {(_date.today().year - dob.year - ((_date.today().month, _date.today().day) < (dob.month, dob.day)))}"
            except Exception:
                pass
        lines.append(f"Athlete: {name_str}{age_str}, gender: {prof_get('gender') or 'not specified'}")
        lines.append(f"Training week: {prof_get('week_start') or 'Saturday'} through {['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'][(['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'].index(prof_get('week_start') or 'Saturday') + 6) % 7]}")
        if prof_get("height_in"):
            h = float(prof_get("height_in"))
            lines.append(f"Height: {int(h//12)}'{int(h%12)}\" ({h} in)")
        if prof_get("target_bw"):
            lines.append(f"Target bodyweight: {prof_get('target_bw')} lbs")
        lines.append("")

    # Add active phase
    active_phase = conn.execute(
        "SELECT phase_type, start_date, label FROM phases WHERE user_id=%s AND end_date IS NULL ORDER BY start_date DESC LIMIT 1",
        (uid,)
    ).fetchone()
    # Add active protocols
    protos = conn.execute(
        "SELECT name, dose, frequency, notes, start_date, end_date FROM protocols WHERE user_id=%s ORDER BY sort_order, id",
        (uid,)
    ).fetchall()
    conn.close()
    if active_phase:
        phase_display = active_phase['phase_type']
        if active_phase.get('label'):
            phase_display += f" ({active_phase['label']})"
        lines.append(f"Current training phase: {phase_display} (since {active_phase['start_date']})")
        lines.append("")

    if protos:
        today = _date.today().isoformat()
        # start_date/end_date are Postgres DATE columns (datetime.date), but TEXT
        # under SQLite. Normalize to ISO strings before comparing to `today` (str),
        # or `date <= str` raises TypeError and the whole endpoint 500s.
        def _iso(x):
            return x.isoformat() if hasattr(x, "isoformat") else x
        active_protos   = [p for p in protos if not p["end_date"] and (not p["start_date"] or _iso(p["start_date"]) <= today)]
        inactive_protos = [p for p in protos if p["end_date"] and _iso(p["end_date"]) < today]
        future_protos   = [p for p in protos if p["start_date"] and _iso(p["start_date"]) > today]

        def proto_line(p):
            parts = [p["name"]]
            if p["dose"]:      parts.append(p["dose"])
            if p["frequency"]: parts.append(p["frequency"])
            if p["notes"]:     parts.append(f"({p['notes']})")
            return "  - " + " | ".join(parts)

        if active_protos:
            lines.append("Active protocols / supplements (currently in use):")
            for p in active_protos:
                line = proto_line(p)
                if p["start_date"]: line += f"  [since {p['start_date']}]"
                lines.append(line)
            lines.append("")

        if inactive_protos:
            lines.append("Past protocols (discontinued):")
            for p in inactive_protos:
                line = proto_line(p)
                line += f"  [{p['start_date'] or '?'} to {p['end_date']}]"
                lines.append(line)
            lines.append("")

        if future_protos:
            lines.append("Upcoming protocols (not yet started):")
            for p in future_protos:
                line = proto_line(p)
                line += f"  [starts {p['start_date']}]"
                lines.append(line)
            lines.append("")

    lines.append(f"Total sessions: {len(rows)}")
    lines.append(f"Date range: {rows[0]['date']} to {rows[-1]['date']}")
    lines.append("")
    lines.append("RD = Relative Density = total session volume / (total time x bodyweight).")
    lines.append("Higher RD means more work done per unit time per pound of bodyweight.")
    lines.append("")
    lines.append("SESSION LOG (date | bodyweight | RD | exercises[load x reps]):")

    for r in rows:
        try:
            exs = r["exercises"] if isinstance(r["exercises"], list) else json.loads(r["exercises"])
        except Exception:
            exs = []
        ex_parts = []
        for ex in exs:
            sets = ex.get("sets", [])
            if not sets:
                continue
            # Compact: name maxload xtotalreps
            max_load = max((s.get("trueLbs") or s.get("rawLoad") or 0) for s in sets)
            total_reps = sum(s.get("reps", 0) for s in sets)
            ex_parts.append(f"{ex.get('name','?')} {max_load:g}x{total_reps:g}")
        note = f" | note: {r['notes']}" if r["notes"] else ""
        sleep_note = ""
        if r.get("sleep_hours") is not None:
            sleep_str = f"{r['sleep_hours']:.1f}h sleep"
            if r.get("deep_sleep_pct") is not None:
                sleep_str += f" ({r['deep_sleep_pct']}% deep)"
            sleep_note = f" | {sleep_str}"
        lines.append(f"{r['date']} | {r['bw']:g}lb | RD {r['rd']:.2f} | {'; '.join(ex_parts)}{note}{sleep_note}")

    return "\n".join(lines)


SYSTEM_PROMPT = """You are an analytical strength-training assistant inside an app called Agon.
The user tracks workouts using a metric called RD (Relative Density). You have access to their
full session history below. Answer their questions directly and concisely, grounding every claim
in the actual data. Use specific dates, loads, and numbers. If the data doesn't support an answer,
say so plainly. Do not invent data. Keep answers tight and useful — this is a knowledgeable user
who wants signal, not filler. When discussing trends, cite the specific sessions that show them."""


def _is_claude_model(model: str) -> bool:
    return model.startswith("claude-")

def _is_ollama_model(model: str) -> bool:
    return model.startswith("ollama:")

async def call_llm(api_key: str, model: str, context: str, history: list, question: str, ollama_url: str = None) -> str:
    """Route to Claude, Gemini, or Ollama depending on the model name."""
    if _is_ollama_model(model):
        if not ollama_url:
            raise HTTPException(status_code=400, detail="Ollama URL not configured")
        model_name = model[7:]  # Strip "ollama:" prefix
        return await _call_ollama(ollama_url, model_name, context, history, question)
    elif _is_claude_model(model):
        return await _call_anthropic(api_key, model, context, history, question)
    return await _call_gemini(api_key, model, context, history, question)

async def _call_ollama(base_url: str, model_name: str, context: str, history: list, question: str) -> str:
    """Call Ollama (OpenAI-compatible API) with training context + conversation history."""
    url = f"{base_url.rstrip('/')}/v1/chat/completions"
    messages = []
    messages.append({
        "role": "system",
        "content": f"{SYSTEM_PROMPT}\n\n=== TRAINING DATA ===\n{context}\n\n=== END DATA ==="
    })
    # Prior conversation turns
    for msg in history:
        messages.append({
            "role": "user" if msg["role"] == "user" else "assistant",
            "content": msg["content"]
        })
    messages.append({"role": "user", "content": question})
    
    payload = {
        "model": model_name,
        "messages": messages,
        "temperature": 0.7,
        "top_p": 0.9,
    }
    
    try:
        async with httpx.AsyncClient(timeout=120) as client:
            resp = await client.post(url, json=payload)
            if resp.status_code != 200:
                raise HTTPException(status_code=502, detail=f"Ollama error ({resp.status_code}): {resp.text[:300]}")
            data = resp.json()
            try:
                return data["choices"][0]["message"]["content"]
            except (KeyError, IndexError):
                raise HTTPException(status_code=502, detail="Ollama returned an unexpected response shape")
    except httpx.ConnectError:
        raise HTTPException(status_code=502, detail=f"Could not reach Ollama at {base_url}. Is it running%s")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Ollama error: {str(e)[:300]}")

async def _call_gemini(api_key: str, model: str, context: str, history: list, question: str) -> str:
    """Call the Gemini API with training context + conversation history."""
    url = f"https://generativelanguage.googleapis.com/v1/models/{model}:generateContent"
    contents = []
    contents.append({
        "role": "user",
        "parts": [{"text": f"{SYSTEM_PROMPT}\n\n=== TRAINING DATA ===\n{context}\n\n=== END DATA ===\n\nAcknowledge you have the data and are ready."}]
    })
    contents.append({
        "role": "model",
        "parts": [{"text": "I have your full training history loaded and I'm ready to analyze it. What would you like to know%s"}]
    })
    for msg in history:
        contents.append({
            "role": "user" if msg["role"] == "user" else "model",
            "parts": [{"text": msg["content"]}]
        })
    contents.append({"role": "user", "parts": [{"text": question}]})
    payload = {"contents": contents}
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(url, params={"key": api_key}, json=payload)
        if resp.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Gemini API error ({resp.status_code}): {resp.text[:300]}")
        data = resp.json()
        try:
            return data["candidates"][0]["content"]["parts"][0]["text"]
        except (KeyError, IndexError):
            raise HTTPException(status_code=502, detail="Gemini returned an unexpected response shape")

async def _call_anthropic(api_key: str, model: str, context: str, history: list, question: str) -> str:
    """Call the Anthropic Messages API with training context + conversation history."""
    url = "https://api.anthropic.com/v1/messages"
    messages = []
    # Prior conversation turns
    for msg in history:
        messages.append({
            "role": "user" if msg["role"] == "user" else "assistant",
            "content": msg["content"]
        })
    messages.append({"role": "user", "content": question})
    payload = {
        "model": model,
        "max_tokens": 1024,
        "system": f"{SYSTEM_PROMPT}\n\n=== TRAINING DATA ===\n{context}\n\n=== END DATA ===",
        "messages": messages
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json"
    }
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(url, headers=headers, json=payload)
        if resp.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Anthropic API error ({resp.status_code}): {resp.text[:300]}")
        data = resp.json()
        try:
            return data["content"][0]["text"]
        except (KeyError, IndexError):
            raise HTTPException(status_code=502, detail="Anthropic returned an unexpected response shape")


@app.get("/insights/history")
def get_insights_history(user=Depends(current_user)):
    """Return persisted conversation. Demo gets nothing (shared identity)."""
    if user["role"] == "demo":
        return {"messages": []}
    conn = get_db()
    rows = conn.execute(
        "SELECT role, content, created_at FROM insights_messages WHERE user_id=%s ORDER BY id",
        (user["id"],)
    ).fetchall()
    conn.close()
    return {"messages": [dict(r) for r in rows]}

@app.delete("/insights/history")
def clear_insights_history(user=Depends(current_user)):
    if user["role"] == "demo":
        return {"status": "noop"}
    conn = get_db()
    conn.execute("DELETE FROM insights_messages WHERE user_id=%s", (user["id"],))
    conn.commit()
    conn.close()
    return {"status": "cleared"}

@app.post("/insights/ask")
async def insights_ask(body: InsightsQuery, user=Depends(current_user)):
    q = (body.question or "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="Empty question")

    # Get the user's key + model + ollama settings
    conn = get_db()
    settings = conn.execute("SELECT ai_key_enc, ai_model, ollama_base_url FROM user_settings WHERE user_id=%s",
                            (user["id"],)).fetchone()
    conn.close()
    
    if not settings or (not settings["ai_key_enc"] and not settings.get("ollama_base_url")):
        raise HTTPException(status_code=400, detail="No AI provider configured. Add one in Settings.")
    
    api_key = decrypt_secret(settings["ai_key_enc"]) if settings["ai_key_enc"] else None
    if settings["ai_key_enc"] and not api_key:
        raise HTTPException(status_code=400, detail="Stored key could not be decrypted. Re-enter it in Settings.")
    
    model = settings["ai_model"] or "gemini-2.5-flash"
    ollama_url = settings.get("ollama_base_url", "")

    # Build context from the data the user is allowed to see
    uid = data_user_id(user)
    context = build_training_context(uid)

    # Load prior conversation (demo has none)
    history = []
    if user["role"] != "demo":
        conn = get_db()
        rows = conn.execute(
            "SELECT role, content FROM insights_messages WHERE user_id=%s ORDER BY id",
            (user["id"],)
        ).fetchall()
        conn.close()
        history = [dict(r) for r in rows]

    answer = await call_llm(api_key, model, context, history, q, ollama_url=ollama_url)

    # Persist both turns (not for demo)
    if user["role"] != "demo":
        conn = get_db()
        conn.execute("INSERT INTO insights_messages (user_id, role, content) VALUES (%s,%s,%s)",
                     (user["id"], "user", q))
        conn.execute("INSERT INTO insights_messages (user_id, role, content) VALUES (%s,%s,%s)",
                     (user["id"], "model", answer))
        conn.commit()
        conn.close()

    log_event(user["id"], user["username"], "insights_ask", f"model={model}")
    return {"answer": answer}

@app.post("/ai/test")
async def test_ai_key(body: AISettingsIn, user=Depends(current_user)):
    """Test a Gemini key with a trivial call. Uses provided key, or stored key if none given."""
    key = (body.ai_key or "").strip()
    model = body.ai_model or "gemini-2.5-flash"
    if not key:
        conn = get_db()
        row = conn.execute("SELECT ai_key_enc FROM user_settings WHERE user_id=%s",
                           (user["id"],)).fetchone()
        conn.close()
        if row and row["ai_key_enc"]:
            key = decrypt_secret(row["ai_key_enc"])
    if not key:
        raise HTTPException(status_code=400, detail="No key to test")

    try:
        if _is_claude_model(model):
            url = "https://api.anthropic.com/v1/messages"
            headers = {"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"}
            payload = {"model": model, "max_tokens": 10, "messages": [{"role": "user", "content": "Reply OK"}]}
            async with httpx.AsyncClient(timeout=20) as client:
                resp = await client.post(url, headers=headers, json=payload)
        else:
            url = f"https://generativelanguage.googleapis.com/v1/models/{model}:generateContent"
            payload = {"contents": [{"role": "user", "parts": [{"text": "Reply with just: OK"}]}]}
            async with httpx.AsyncClient(timeout=20) as client:
                resp = await client.post(url, params={"key": key}, json=payload)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Connection failed: {str(e)[:200]}")
    if resp.status_code == 200:
        return {"ok": True, "model": model}
    raise HTTPException(status_code=400, detail=f"Key test failed ({resp.status_code}): {resp.text[:200]}")

# ── Import ────────────────────────────────────────────────────
from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@app.get("/import/template")
def download_import_template(user=Depends(current_user)):
    """Generate and return a filled-out XLSX import template."""
    wb = openpyxl.Workbook()

    # ── Log sheet ──
    ws = wb.active
    ws.title = "Log"

    headers = ["Date", "Bodyweight (lbs)", "Exercise", "Tool", "Modifier", "Load", "Reps", "Time (min)", "Phase"]
    header_fill = PatternFill("solid", fgColor="1A1A18")
    header_font = Font(bold=True, color="E05A20")
    border = Border(bottom=Side(style="thin", color="444444"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Column widths
    widths = [14, 18, 28, 14, 20, 10, 8, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Example rows
    examples = [
        ["01/15/2026", 200.0, "Bench Press",     "Bar",        "Standard",  185, 5,  1.5],
        ["01/15/2026", 200.0, "Bench Press",     "Bar",        "Standard",  185, 5,  1.5],
        ["01/15/2026", 200.0, "Bench Press",     "Bar",        "Standard",  185, 5,  1.5],
        ["01/15/2026", 200.0, "Incline Fly",     "Cable",      "Double",    40,  12, 1.5],
        ["01/15/2026", 200.0, "Incline Fly",     "Cable",      "Double",    40,  12, 1.5],
        ["01/15/2026", 200.0, "Pull-up",         "Bodyweight", "Vest",      25,  8,  1.5],
        ["01/15/2026", 200.0, "Pull-up",         "Bodyweight", "Vest",      25,  8,  1.5],
        ["01/17/2026", 199.5, "Low Bar Squat",   "Bar",        "Standard",  225, 3,  2.0],
        ["01/17/2026", 199.5, "KB Swings",       "Kettlebell", "",          53,  15, 1.5, "Recomp"],
    ]
    example_font = Font(color="888888", italic=True)
    for r, row in enumerate(examples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = example_font

    ws.freeze_panes = "A2"

    # ── Legend sheet ──
    wl = wb.create_sheet("Legend")
    wl.column_dimensions["A"].width = 20
    wl.column_dimensions["B"].width = 60
    wl.column_dimensions["C"].width = 50

    legend_title_font = Font(bold=True, color="E05A20", size=13)
    section_font = Font(bold=True, color="CCCCCC")
    body_font = Font(color="AAAAAA")
    wl["A1"] = "Agon Import — Field Legend"
    wl["A1"].font = legend_title_font

    rows = [
        ("", "", ""),
        ("FIELD", "DESCRIPTION", "VALID VALUES / NOTES"),
        ("Date", "Session date", "MM/DD/YYYY — all sets for a session share the same date"),
        ("Bodyweight (lbs)", "Your bodyweight that day", "Decimal OK — e.g. 199.5"),
        ("Exercise", "Exercise name", "Any name. New exercises will be created automatically."),
        ("Tool", "Equipment type", "Bar | Cable | Kettlebell | Dumbbell | Machine | Bodyweight"),
        ("Modifier", "Sub-type of tool or bodyweight mode", "See below"),
        ("Load", "Weight entered (see modifier rules)", "Numeric. See below for what to enter per modifier."),
        ("Reps", "Reps performed in this set", "Integer"),
        ("Time (min)", "Duration of this set in minutes", "Decimal. e.g. 1.5 for 90 seconds. Defaults to 1.5 if left blank."),
        ("Phase", "Training phase active on this date (optional)", "Weight Loss | Cut | Recomp | Bulk | Maintain | Peak | Deload. Leave blank to skip. Consecutive rows with the same phase are grouped into one phase record."),
        ("", "", ""),
        ("MODIFIER GUIDE", "", ""),
        ("Tool: Bar", "", ""),
        ("  Standard", "Two-sided barbell, enter total weight", "Load = total bar weight (e.g. 135 for 45lb bar + 45lb per side)"),
        ("  Vitruvian", "Vitruvian bar, enter weight per side", "Load = plates per side (trueLbs = Load × 2)"),
        ("Tool: Cable", "", ""),
        ("  Single", "One cable / one arm", "Load = stack weight (trueLbs = Load × 1)"),
        ("  Double", "Two cables / both arms", "Load = stack weight per side (trueLbs = Load × 2)"),
        ("Tool: Dumbbell", "", ""),
        ("  (leave blank)", "Always counts both dumbbells", "Load = weight of one dumbbell (trueLbs = Load × 2)"),
        ("Tool: Kettlebell", "", ""),
        ("  (leave blank)", "Single bell", "Load = bell weight (trueLbs = Load × 1)"),
        ("Tool: Machine", "", ""),
        ("  (leave blank)", "Machine stack", "Load = stack weight (trueLbs = Load × 1)"),
        ("Tool: Bodyweight", "", ""),
        ("  Normal", "Just bodyweight", "Load = 0 or leave blank (trueLbs = bodyweight)"),
        ("  Vest", "Bodyweight + weighted vest", "Load = vest weight in lbs (trueLbs = bodyweight + vest)"),
        ("  Assisted", "Bodyweight with band/machine assist", "Load = assist % (e.g. 20 = 20% assist, trueLbs = bodyweight × 0.80)"),
        ("", "", ""),
        ("NOTES", "", ""),
        ("", "Each ROW is one SET.", "Group sets for the same exercise on the same date together."),
        ("", "Time per set defaults to 1.5 min.", "All imported sets use 1.5 min for RD calculation."),
        ("", "Existing sessions on the same date will NOT be overwritten.", "Change the date or delete the existing session first."),
        ("", "Grey rows in the Log sheet are examples — delete them before importing.", ""),
    ]

    for r, (a, b, c) in enumerate(rows, 2):
        wl.cell(row=r, column=1, value=a)
        wl.cell(row=r, column=2, value=b)
        wl.cell(row=r, column=3, value=c)
        if a in ("FIELD", "MODIFIER GUIDE", "NOTES"):
            for col in range(1, 4):
                wl.cell(row=r, column=col).font = section_font
        elif a in ("Tool: Bar", "Tool: Cable", "Tool: Dumbbell", "Tool: Kettlebell", "Tool: Machine", "Tool: Bodyweight"):
            for col in range(1, 4):
                wl.cell(row=r, column=col).font = Font(bold=True, color="E05A20")
        else:
            for col in range(1, 4):
                wl.cell(row=r, column=col).font = body_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=agon_import_template.xlsx"}
    )


def _parse_import_row(row, bw: float) -> dict:
    """Convert a single spreadsheet row into a set dict with optional phase."""
    tool      = (str(row[3] or "Bar")).strip()
    modifier  = (str(row[4] or "")).strip().lower()
    raw_load  = float(row[5] or 0)
    reps      = int(float(row[6] or 0))
    time_min  = float(row[7]) if len(row) > 7 and row[7] not in (None, "", " ") else 1.5
    phase_raw = row[8] if len(row) > 8 else None
    phase     = str(phase_raw).strip() if phase_raw not in (None, "", " ", "None") else None

    # Determine trueLbs based on tool + modifier
    if tool == "Bar":
        if "vitruvian" in modifier:
            true_lbs = raw_load * 2
            mult = 2
        else:  # Standard
            true_lbs = raw_load
            mult = 2
    elif tool == "Cable":
        if "double" in modifier:
            true_lbs = raw_load * 2
            mult = 2
        else:  # Single
            true_lbs = raw_load
            mult = 1
    elif tool == "Dumbbell":
        true_lbs = raw_load * 2
        mult = 2
    elif tool == "Kettlebell":
        true_lbs = raw_load
        mult = 1
    elif tool == "Machine":
        true_lbs = raw_load
        mult = 1
    elif tool == "Bodyweight":
        if "vest" in modifier:
            true_lbs = bw + raw_load
        elif "assisted" in modifier or "assist" in modifier:
            true_lbs = bw * (1 - raw_load / 100.0)
        else:
            true_lbs = bw
        mult = 1
    else:
        true_lbs = raw_load
        mult = 1

    vol = round(true_lbs * reps, 2)
    return {
        "rawLoad": raw_load,
        "trueLbs": round(true_lbs, 2),
        "reps": reps,
        "vol": vol,
        "time":  time_min,
        "phase": phase
    }


@app.post("/import/sessions")
async def import_sessions(user=Depends(current_user), file: UploadFile = File(...)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot import data")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read XLSX: {str(e)[:200]}")

    # Parse rows (skip header row 1)
    from collections import defaultdict, OrderedDict
    sessions_map = OrderedDict()  # key: (date, bw) → {exercises: {name → {tool, sets[]}}}

    skipped = []
    errors  = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            raw_date = str(row[0] or "").strip()
            bw       = float(row[1] or 0)
            ex_name  = str(row[2] or "").strip()
            tool     = str(row[3] or "Bar").strip()
            modifier = str(row[4] or "").strip()

            if not raw_date or not ex_name or not bw:
                continue

            # Normalise date to YYYY-MM-DD
            from datetime import datetime
            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%d/%m/%Y"):
                try:
                    date_str = datetime.strptime(raw_date, fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue
            else:
                errors.append(f"Row {row_idx}: unrecognised date format '{raw_date}'")
                continue

            set_data = _parse_import_row(row, bw)
            if set_data["reps"] <= 0:
                continue

            key = (date_str, bw)
            if key not in sessions_map:
                sessions_map[key] = {}
            if ex_name not in sessions_map[key]:
                sessions_map[key][ex_name] = {"tool": tool, "modifier": modifier, "sets": []}
            sessions_map[key][ex_name]["sets"].append(set_data)

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)[:100]}")

    if not sessions_map:
        raise HTTPException(status_code=400, detail="No valid rows found. Check the template format.")

    # Build session objects and insert
    conn = get_db()
    inserted = 0
    skipped_dates = []

    for (date_str, bw), exercises_map in sessions_map.items():
        # Check for existing session
        existing = conn.execute(
            "SELECT id FROM sessions WHERE user_id=%s AND date=%s",
            (user["id"], date_str)
        ).fetchone()
        if existing:
            skipped_dates.append(date_str)
            continue

        exercises = []
        for ex_name, ex_data in exercises_map.items():
            sets = ex_data["sets"]
            total_vol  = sum(s["vol"]  for s in sets)
            total_time = sum(s["time"] for s in sets)
            density = round(total_vol / total_time, 2) if total_time > 0 else 0
            exercises.append({
                "name":     ex_name,
                "tool":     ex_data["tool"],
                "sets":     sets,
                "totalVol": round(total_vol, 2),
                "density":  density,
                "isBW":     ex_data["tool"] == "Bodyweight"
            })

        total_vol  = sum(s["vol"]  for ex in exercises for s in ex["sets"])
        total_time = sum(s["time"] for ex in exercises for s in ex["sets"])
        rd = round(total_vol / (total_time * bw), 4) if (total_time > 0 and bw) else 0
        total_density = round(total_vol / total_time, 4) if total_time > 0 else 0

        conn.execute(
            "INSERT INTO sessions (user_id, date, bw, rd, total_density, exercises) VALUES (%s,%s,%s,%s,%s,%s)",
            (user["id"], date_str, bw, rd, total_density, json.dumps(exercises))
        )
        inserted += 1

    # Synthesize phases from Phase column — group consecutive same-phase dates
    date_phase_map = {}  # date -> phase_type
    for (d_str, bw_val), ex_map in sessions_map.items():
        for ex_name, ex_data in ex_map.items():
            for s in ex_data["sets"]:
                if s.get("phase"):
                    date_phase_map[d_str] = s["phase"]
                    break
            break

    phases_created = 0
    if date_phase_map:
        sorted_dates = sorted(date_phase_map.keys())
        runs = []  # [(phase_type, start, end)]
        cur_type  = date_phase_map[sorted_dates[0]]
        cur_start = sorted_dates[0]
        prev_date = sorted_dates[0]
        for d in sorted_dates[1:]:
            ptype = date_phase_map[d]
            if ptype != cur_type:
                runs.append((cur_type, cur_start, prev_date))
                cur_type  = ptype
                cur_start = d
            prev_date = d
        runs.append((cur_type, cur_start, None))  # last run stays open

        for phase_type, start_date, end_date in runs:
            if not phase_type:
                continue
            existing = conn.execute(
                "SELECT id FROM phases WHERE user_id=%s AND phase_type=%s AND start_date=%s",
                (user["id"], phase_type, start_date)
            ).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO phases (user_id, phase_type, start_date, end_date) VALUES (%s,%s,%s,%s)",
                    (user["id"], phase_type, start_date, end_date)
                )
                phases_created += 1

    conn.commit()
    conn.close()

    log_event(user["id"], user["username"], "import",
              f"inserted={inserted} phases={phases_created} skipped={len(skipped_dates)}")
    return {
        "inserted":       inserted,
        "skipped":        skipped_dates,
        "errors":         errors,
        "phases_created": phases_created,
    }

# ── Admin: event log ─────────────────────────────────────────
@app.get("/admin/events")
def get_events(user=Depends(current_user), limit: int = 200):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    conn = get_db()
    rows = conn.execute(
        """SELECT e.created_at, e.username, e.event_type, e.detail, e.ip
           FROM events e ORDER BY e.id DESC LIMIT %s""",
        (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/admin/events/summary")
def get_events_summary(user=Depends(current_user)):
    """Per-user activity summary for admin."""
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    conn = get_db()
    rows = conn.execute("""
        SELECT
            username,
            COUNT(*) as total_events,
            SUM(CASE WHEN event_type='login' THEN 1 ELSE 0 END) as logins,
            SUM(CASE WHEN event_type='session_saved' THEN 1 ELSE 0 END) as sessions_saved,
            SUM(CASE WHEN event_type='insights_ask' THEN 1 ELSE 0 END) as insights_queries,
            SUM(CASE WHEN event_type='import' THEN 1 ELSE 0 END) as imports,
            MAX(created_at) as last_seen
        FROM events
        GROUP BY username
        ORDER BY last_seen DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── Phase endpoints ───────────────────────────────────────────
@app.get("/phases")
def get_phases(user=Depends(current_user)):
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT id, phase_type, start_date, end_date, notes, label FROM phases WHERE user_id=%s ORDER BY start_date DESC",
        (uid,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/phases")
def create_phase(body: PhaseIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot create phases")
    from datetime import date as _date
    today_str = _date.today().strftime("%Y-%m-%d")
    conn = get_db()
    # Only auto-close the open phase if new phase starts today or in future
    # For historical phases (past start date) the user sets end dates manually
    if not body.end_date and body.start_date >= today_str:
        conn.execute(
            "UPDATE phases SET end_date=%s WHERE user_id=%s AND end_date IS NULL",
            (body.start_date, user["id"])
        )
    conn.execute(
        "INSERT INTO phases (user_id, phase_type, start_date, end_date, notes, label) VALUES (%s,%s,%s,%s,%s,%s)",
        (user["id"], body.phase_type, body.start_date, body.end_date or None, body.notes or None, body.label or None)
    )
    conn.commit()
    conn.close()
    return {"status": "created"}

@app.delete("/phases/{phase_id}")
def delete_phase(phase_id: int, user=Depends(current_user)):
    conn = get_db()
    conn.execute("DELETE FROM phases WHERE id=%s AND user_id=%s", (phase_id, user["id"]))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

@app.put("/phases/{phase_id}/close")
def close_phase(phase_id: int, user=Depends(current_user)):
    today = _date.today().strftime("%Y-%m-%d")
    conn = get_db()
    conn.execute(
        "UPDATE phases SET end_date=%s WHERE id=%s AND user_id=%s AND end_date IS NULL",
        (today, phase_id, user["id"])
    )
    conn.commit()
    conn.close()
    return {"status": "closed"}


# ── Recomposition score ───────────────────────────────────────
@app.get("/recomp")
def get_recomp(user=Depends(current_user)):
    """
    Recomp score: measures whether RD is rising while BW is falling.
    Uses last 8 weeks of data. Returns a score and trend interpretation.
    Score: positive = gaining strength while losing weight (ideal recomp)
           near zero = maintaining
           negative  = losing strength relative to bodyweight
    """
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT date, bw, rd FROM sessions WHERE user_id=%s ORDER BY date DESC LIMIT 60",
        (uid,)
    ).fetchall()
    # Get target_bw — defensive in case column not yet migrated
    try:
        settings = conn.execute("SELECT target_bw FROM user_settings WHERE user_id=%s", (uid,)).fetchone()
    except Exception:
        settings = None
    conn.close()

    if len(rows) < 4:
        return {"score": None, "label": "Not enough data", "bw_trend": None, "rd_trend": None, "target_bw": settings["target_bw"] if settings else None}

    rows = list(reversed(rows))  # chronological
    # Split into two halves
    mid = len(rows) // 2
    first_half  = rows[:mid]
    second_half = rows[mid:]

    avg_rd_early = sum(r["rd"] for r in first_half)  / len(first_half)
    avg_rd_late  = sum(r["rd"] for r in second_half) / len(second_half)
    avg_bw_early = sum(r["bw"] for r in first_half)  / len(first_half)
    avg_bw_late  = sum(r["bw"] for r in second_half) / len(second_half)

    bw_trend = round(avg_bw_late - avg_bw_early, 1)   # negative = losing weight
    rd_trend = round(avg_rd_late - avg_rd_early, 3)    # positive = getting stronger

    # Score: RD change normalised, penalised if BW is rising (bulk is fine but label differently)
    score = round(rd_trend * 100, 1)

    if rd_trend > 0.02 and bw_trend < 0:
        label = "Recomping"
    elif rd_trend > 0.02 and bw_trend >= 0:
        label = "Building"
    elif abs(rd_trend) <= 0.02 and bw_trend < -0.5:
        label = "Cutting"
    elif abs(rd_trend) <= 0.02:
        label = "Maintaining"
    else:
        label = "Regressing"

    # Current bw vs target
    current_bw = rows[-1]["bw"] if rows else None
    target_bw  = settings["target_bw"] if settings else None
    to_goal    = round(current_bw - target_bw, 1) if (current_bw and target_bw) else None

    return {
        "score":      score,
        "label":      label,
        "bw_trend":   bw_trend,
        "rd_trend":   rd_trend,
        "current_bw": current_bw,
        "target_bw":  target_bw,
        "to_goal":    to_goal,
    }

# ── Protocol endpoints ────────────────────────────────────────
@app.get("/protocols")
def get_protocols(user=Depends(current_user)):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, name, dose, frequency, notes, start_date, end_date, track FROM protocols WHERE user_id=%s ORDER BY CASE WHEN end_date IS NULL THEN 0 ELSE 1 END ASC, end_date DESC, start_date DESC",
        (user["id"],)
    ).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d["track"] = bool(d.get("track", False))
        result.append(d)
    return result

@app.post("/protocols")
def add_protocol(body: ProtocolIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot add protocols")
    try:
        conn = get_db()
        row = conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) AS max_order FROM protocols WHERE user_id=%s", (user["id"],)
        ).fetchone()
        max_order = row["max_order"] if row else 0
        # Coerce empty strings to None for optional date fields
        start_date = body.start_date if body.start_date else None
        end_date = body.end_date if body.end_date else None
        conn.execute(
            "INSERT INTO protocols (user_id, name, dose, frequency, notes, start_date, end_date, sort_order, track) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (user["id"], body.name.strip(), body.dose, body.frequency, body.notes, start_date, end_date, max_order + 1, bool(body.track))
        )
        conn.commit()
        conn.close()
        return {"status": "added"}
    except Exception as e:
        import traceback
        print(f"ERROR in add_protocol: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Protocol add failed: {str(e)}")

@app.put("/protocols/{protocol_id}")
def update_protocol(protocol_id: int, body: ProtocolIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot modify protocols")
    try:
        conn = get_db()
        start_date = body.start_date if body.start_date else None
        end_date = body.end_date if body.end_date else None
        conn.execute(
            "UPDATE protocols SET name=%s, dose=%s, frequency=%s, notes=%s, start_date=%s, end_date=%s, track=%s WHERE id=%s AND user_id=%s",
            (body.name.strip(), body.dose, body.frequency, body.notes, start_date, end_date, bool(body.track), protocol_id, user["id"])
        )
        conn.commit()
        conn.close()
        return {"status": "updated"}
    except Exception as e:
        import traceback
        print(f"ERROR in update_protocol: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Protocol update failed: {str(e)}")

@app.patch("/protocols/{protocol_id}")
def patch_protocol(protocol_id: int, body: ProtocolTrack, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot modify protocols")
    fields, params = [], []
    if body.track is not None:
        fields.append("track=%s"); params.append(bool(body.track))
    if body.end_date is not None:
        fields.append("end_date=%s"); params.append(body.end_date)
    if not fields:
        return {"status": "nothing to update"}
    params.extend([protocol_id, user["id"]])
    conn = get_db()
    conn.execute(f"UPDATE protocols SET {', '.join(fields)} WHERE id=%s AND user_id=%s", params)
    conn.commit()
    conn.close()
    return {"status": "updated"}

@app.delete("/protocols/{protocol_id}")
def delete_protocol(protocol_id: int, user=Depends(current_user)):
    conn = get_db()
    conn.execute("DELETE FROM protocols WHERE id=%s AND user_id=%s", (protocol_id, user["id"]))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

# ── Inventory & dose tracking ─────────────────────────────────

def _item_extras(d):
    """Computed fields: reconstitution math + doses remaining."""
    try:
        if d.get("bac_water_ml") and d.get("total_amount") and d.get("unit") == "mg":
            total_units = float(d["bac_water_ml"]) * 100  # U-100 syringe units
            d["mcg_per_unit"] = round(float(d["total_amount"]) * 1000 / total_units, 2)
            if d.get("per_dose"):
                d["units_per_dose"] = round(float(d["per_dose"]) * 1000 / d["mcg_per_unit"], 1)
        if d.get("per_dose") and d.get("remaining") is not None and float(d["per_dose"]) > 0:
            d["doses_remaining"] = int(float(d["remaining"]) // float(d["per_dose"]))
    except (TypeError, ValueError, ZeroDivisionError):
        pass
    return d

@app.get("/inventory")
def get_inventory(include_empty: bool = False, user=Depends(current_user)):
    conn = get_db()
    sql = """SELECT id, protocol_id, name, item_type, total_amount, unit, bac_water_ml,
                    remaining, per_dose, opened_date, status, vendor, cost, notes, created_at
             FROM inventory_items WHERE user_id=%s"""
    if not include_empty:
        sql += " AND status != 'empty'"
    sql += " ORDER BY CASE status WHEN 'active' THEN 0 WHEN 'sealed' THEN 1 ELSE 2 END, name"
    rows = conn.execute(sql, (user["id"],)).fetchall()
    conn.close()
    return [_item_extras(dict(r)) for r in rows]

@app.post("/inventory")
def add_inventory_item(body: InventoryItemIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot modify inventory")
    remaining = body.remaining if body.remaining is not None else body.total_amount
    conn = get_db()
    conn.execute(
        """INSERT INTO inventory_items
           (user_id, protocol_id, name, item_type, total_amount, unit, bac_water_ml,
            remaining, per_dose, opened_date, status, vendor, cost, notes)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (user["id"], body.protocol_id, body.name.strip(), body.item_type, body.total_amount,
         body.unit, body.bac_water_ml, remaining, body.per_dose,
         body.opened_date or None, body.status, body.vendor, body.cost, body.notes)
    )
    conn.commit()
    conn.close()
    return {"status": "added"}

@app.put("/inventory/{item_id}")
def update_inventory_item(item_id: int, body: InventoryItemIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot modify inventory")
    conn = get_db()
    conn.execute(
        """UPDATE inventory_items SET protocol_id=%s, name=%s, item_type=%s, total_amount=%s,
           unit=%s, bac_water_ml=%s, remaining=%s, per_dose=%s, opened_date=%s, status=%s,
           vendor=%s, cost=%s, notes=%s
           WHERE id=%s AND user_id=%s""",
        (body.protocol_id, body.name.strip(), body.item_type, body.total_amount, body.unit,
         body.bac_water_ml, body.remaining, body.per_dose, body.opened_date or None,
         body.status, body.vendor, body.cost, body.notes, item_id, user["id"])
    )
    conn.commit()
    conn.close()
    return {"status": "updated"}

@app.delete("/inventory/{item_id}")
def delete_inventory_item(item_id: int, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot modify inventory")
    conn = get_db()
    conn.execute("DELETE FROM inventory_items WHERE id=%s AND user_id=%s", (item_id, user["id"]))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

@app.get("/doses")
def get_doses(date_from: Optional[str] = None, date_to: Optional[str] = None,
              protocol_id: Optional[int] = None, limit: int = 200,
              user=Depends(current_user)):
    conn = get_db()
    sql = """SELECT d.id, d.protocol_id, d.item_id, d.amount, d.unit, d.taken_at,
                    d.site, d.notes, p.name AS protocol_name, i.name AS item_name
             FROM dose_events d
             LEFT JOIN protocols p ON p.id = d.protocol_id
             LEFT JOIN inventory_items i ON i.id = d.item_id
             WHERE d.user_id=%s"""
    params = [user["id"]]
    if date_from:
        sql += " AND d.taken_at >= %s"; params.append(date_from)
    if date_to:
        sql += " AND d.taken_at < (%s::date + INTERVAL '1 day')"; params.append(date_to)
    if protocol_id:
        sql += " AND d.protocol_id = %s"; params.append(protocol_id)
    sql += " ORDER BY d.taken_at DESC LIMIT %s"; params.append(min(limit, 1000))
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/doses")
def log_dose(body: DoseEventIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot log doses")
    conn = get_db()
    try:
        amount, unit, low_stock = body.amount, body.unit, None
        if body.item_id:
            item = conn.execute(
                "SELECT remaining, per_dose, unit, status FROM inventory_items WHERE id=%s AND user_id=%s",
                (body.item_id, user["id"])
            ).fetchone()
            if not item:
                raise HTTPException(status_code=404, detail="Inventory item not found")
            if amount is None:
                amount = float(item["per_dose"]) if item["per_dose"] is not None else None
            if unit is None:
                unit = item["unit"]
            if amount is not None and item["remaining"] is not None:
                new_remaining = max(float(item["remaining"]) - amount, 0)
                new_status = "empty" if new_remaining <= 0 else ("active" if item["status"] == "sealed" else item["status"])
                conn.execute(
                    "UPDATE inventory_items SET remaining=%s, status=%s WHERE id=%s AND user_id=%s",
                    (new_remaining, new_status, body.item_id, user["id"])
                )
                per_dose = float(item["per_dose"]) if item["per_dose"] else None
                if per_dose and new_remaining / per_dose <= 5:
                    low_stock = {"remaining": new_remaining, "doses_remaining": int(new_remaining // per_dose)}
        conn.execute(
            """INSERT INTO dose_events (user_id, protocol_id, item_id, amount, unit, taken_at, site, notes)
               VALUES (%s,%s,%s,%s,%s,COALESCE(%s, CURRENT_TIMESTAMP),%s,%s)""",
            (user["id"], body.protocol_id, body.item_id, amount, unit,
             body.taken_at or None, body.site, body.notes)
        )
        conn.commit()
        resp = {"status": "logged"}
        if low_stock:
            resp["low_stock"] = low_stock
        return resp
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Dose log failed: {str(e)}")
    finally:
        conn.close()

@app.delete("/doses/{dose_id}")
def delete_dose(dose_id: int, restock: bool = True, user=Depends(current_user)):
    """Delete a dose event; by default restores the amount to the linked item."""
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT item_id, amount FROM dose_events WHERE id=%s AND user_id=%s",
            (dose_id, user["id"])
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Dose event not found")
        if restock and row["item_id"] and row["amount"] is not None:
            conn.execute(
                """UPDATE inventory_items
                   SET remaining = LEAST(COALESCE(remaining,0) + %s, COALESCE(total_amount, COALESCE(remaining,0) + %s)),
                       status = CASE WHEN status='empty' THEN 'active' ELSE status END
                   WHERE id=%s AND user_id=%s""",
                (row["amount"], row["amount"], row["item_id"], user["id"])
            )
        conn.execute("DELETE FROM dose_events WHERE id=%s AND user_id=%s", (dose_id, user["id"]))
        conn.commit()
        return {"status": "deleted"}
    except HTTPException:
        conn.rollback()
        raise
    finally:
        conn.close()

# ── External API key management ───────────────────────────────
import secrets as _secrets

@app.get("/external/key")
def get_external_key(user=Depends(current_user)):
    """Return masked external API key status."""
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot have external API keys")
    conn = get_db()
    try:
        row = conn.execute("SELECT external_api_key FROM user_settings WHERE user_id=%s", (user["id"],)).fetchone()
    except Exception:
        row = None
    conn.close()
    key = row["external_api_key"] if row and row["external_api_key"] else None
    if not key:
        return {"has_key": False, "masked": ""}
    return {"has_key": True, "masked": key[:8] + "..." + key[-4:]}

@app.post("/external/key/generate")
def generate_external_key(user=Depends(current_user)):
    """Generate a new external API key for the user."""
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot have external API keys")
    new_key = "agon_" + _secrets.token_urlsafe(32)
    conn = get_db()
    existing = conn.execute("SELECT user_id FROM user_settings WHERE user_id=%s", (user["id"],)).fetchone()
    if existing:
        conn.execute("UPDATE user_settings SET external_api_key=%s WHERE user_id=%s", (new_key, user["id"]))
    else:
        conn.execute("INSERT INTO user_settings (user_id, external_api_key) VALUES (%s,%s)", (user["id"], new_key))
    conn.commit()
    conn.close()
    return {"key": new_key}

@app.delete("/external/key")
def revoke_external_key(user=Depends(current_user)):
    conn = get_db()
    conn.execute("UPDATE user_settings SET external_api_key=NULL WHERE user_id=%s", (user["id"],))
    conn.commit()
    conn.close()
    return {"status": "revoked"}

def _resolve_external_key(authorization, x_api_key):
    """Resolve an external API key (Bearer or X-API-Key) to a user row, or raise 401."""
    api_key = None
    if authorization and authorization.lower().startswith("bearer "):
        api_key = authorization[7:].strip()
    elif x_api_key:
        api_key = x_api_key.strip()
    if not api_key:
        raise HTTPException(status_code=401, detail="API key required (use Authorization: Bearer header)")
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT u.id, u.username FROM users u "
            "JOIN user_settings us ON us.user_id = u.id "
            "WHERE us.external_api_key=%s", (api_key,)
        ).fetchone()
    except Exception:
        row = None
    conn.close()
    if not row:
        raise HTTPException(status_code=401, detail="Invalid or revoked API key")
    return dict(row)

class ExternalDoseIn(BaseModel):
    compound:    str                          # free-text compound name (log is self-describing)
    amount:      Optional[float] = None
    unit:        Optional[str] = None
    taken_at:    Optional[str] = None         # ISO datetime; defaults to now
    site:        Optional[str] = None
    notes:       Optional[str] = None
    protocol_id: Optional[int] = None
    item_id:     Optional[int] = None         # if set, decrements inventory (live logging)
    status:      Optional[str] = "Logged"

@app.post("/api/external/doses")
def external_log_dose(body: ExternalDoseIn,
                      x_api_key: Optional[str] = Header(default=None),
                      authorization: Optional[str] = Header(default=None)):
    """Write a single dose via external API key (for the Telegram bot / automations).
    If item_id is supplied, inventory is decremented and a low-stock flag returned."""
    ext = _resolve_external_key(authorization, x_api_key)
    uid = ext["id"]
    conn = get_db()
    try:
        amount, unit, low_stock = body.amount, body.unit, None
        if body.item_id:
            item = conn.execute(
                "SELECT remaining, per_dose, unit, status FROM inventory_items WHERE id=%s AND user_id=%s",
                (body.item_id, uid)
            ).fetchone()
            if item:
                if amount is None and item["per_dose"] is not None:
                    amount = float(item["per_dose"])
                if unit is None:
                    unit = item["unit"]
                if amount is not None and item["remaining"] is not None:
                    new_rem = max(float(item["remaining"]) - amount, 0)
                    new_status = "empty" if new_rem <= 0 else ("active" if item["status"] == "sealed" else item["status"])
                    conn.execute("UPDATE inventory_items SET remaining=%s, status=%s WHERE id=%s AND user_id=%s",
                                 (new_rem, new_status, body.item_id, uid))
                    pd = float(item["per_dose"]) if item["per_dose"] else None
                    if pd and new_rem / pd <= 5:
                        low_stock = {"remaining": new_rem, "doses_remaining": int(new_rem // pd)}
        note = body.notes
        if body.compound:
            tag = f"[{body.compound}]"
            note = f"{tag} {note}" if note else tag
        conn.execute(
            """INSERT INTO dose_events (user_id, protocol_id, item_id, amount, unit, taken_at, site, notes)
               VALUES (%s,%s,%s,%s,%s,COALESCE(%s, CURRENT_TIMESTAMP),%s,%s)""",
            (uid, body.protocol_id, body.item_id, amount, unit, body.taken_at or None, body.site, note)
        )
        conn.commit()
        resp = {"status": "logged"}
        if low_stock:
            resp["low_stock"] = low_stock
        return resp
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"External dose log failed: {str(e)}")
    finally:
        conn.close()

@app.post("/api/external/doses/import")
def external_import_doses(payload: dict,
                          x_api_key: Optional[str] = Header(default=None),
                          authorization: Optional[str] = Header(default=None)):
    """Bulk historical dose import via external API key. Body: {"doses": [ExternalDoseIn...]}.
    Historical rows never touch inventory (item_id ignored) — stock is counted fresh.
    Auto-links protocol_id by exact protocol-name match when a 'protocol' field is given.
    Idempotent per (taken_at, compound, amount) so re-running won't duplicate."""
    ext = _resolve_external_key(authorization, x_api_key)
    uid = ext["id"]
    doses = payload.get("doses", [])
    if not isinstance(doses, list):
        raise HTTPException(status_code=400, detail="Body must be {'doses': [...]}")
    conn = get_db()
    try:
        # protocol name -> id map for linking
        prows = conn.execute("SELECT id, name FROM protocols WHERE user_id=%s", (uid,)).fetchall()
        pmap = { (r["name"] or "").strip().lower(): r["id"] for r in prows }
        inserted, skipped, linked = 0, 0, 0
        for d in doses:
            compound = (d.get("compound") or "").strip()
            if not compound:
                skipped += 1; continue
            taken_at = d.get("taken_at")
            amount = d.get("amount")
            # idempotency guard
            existing = conn.execute(
                """SELECT 1 FROM dose_events
                   WHERE user_id=%s AND taken_at=%s AND COALESCE(notes,'') LIKE %s
                   LIMIT 1""",
                (uid, taken_at, f"[{compound}]%")
            ).fetchone()
            if existing:
                skipped += 1; continue
            pid = d.get("protocol_id")
            pname = (d.get("protocol") or "").strip().lower()
            if pid is None and pname and pname in pmap:
                pid = pmap[pname]; linked += 1
            note = d.get("notes")
            tag = f"[{compound}]"
            if d.get("status") and d.get("status") != "Logged":
                tag += f"[{d.get('status')}]"
            note = f"{tag} {note}" if note else tag
            conn.execute(
                """INSERT INTO dose_events (user_id, protocol_id, item_id, amount, unit, taken_at, site, notes)
                   VALUES (%s,%s,NULL,%s,%s,%s,%s,%s)""",
                (uid, pid, amount, d.get("unit"), taken_at, d.get("site"), note)
            )
            inserted += 1
        conn.commit()
        return {"status": "ok", "inserted": inserted, "skipped_existing": skipped, "protocol_linked": linked}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")
    finally:
        conn.close()

@app.get("/external/data")
def get_external_data(request: Request,
                      x_api_key: Optional[str] = Header(default=None),
                      authorization: Optional[str] = Header(default=None)):
    """Public read-only endpoint for external consumers.
    Authenticated by external API key, supplied via (in order of preference):
      - Authorization: Bearer <key>
      - X-API-Key: <key>
    The key is never accepted as a query parameter, to keep it out of access logs."""
    api_key = None
    if authorization and authorization.lower().startswith("bearer "):
        api_key = authorization[7:].strip()
    elif x_api_key:
        api_key = x_api_key.strip()

    if not api_key:
        raise HTTPException(status_code=401, detail="API key required (use Authorization: Bearer header)")

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT u.id, u.username, us.external_api_key FROM users u "
            "JOIN user_settings us ON us.user_id = u.id "
            "WHERE us.external_api_key=%s", (api_key,)
        ).fetchone()
    except Exception:
        row = None
    if not row:
        conn.close()
        raise HTTPException(status_code=401, detail="Invalid or revoked API key")

    uid = row["id"]
    username = row["username"]

    # Sessions
    sessions = conn.execute(
        "SELECT date, bw, rd, total_density, exercises, notes, sleep_hours, deep_sleep_pct FROM sessions WHERE user_id=%s ORDER BY date",
        (uid,)
    ).fetchall()

    # Profile
    try:
        prof = conn.execute(
            "SELECT first_name, last_name, dob, gender, week_start, height_in, target_bw, activity_level FROM user_settings WHERE user_id=%s",
            (uid,)
        ).fetchone()
    except Exception:
        prof = None

    # Phases
    phases = conn.execute(
        "SELECT phase_type, start_date, end_date, notes FROM phases WHERE user_id=%s ORDER BY start_date",
        (uid,)
    ).fetchall()

    # Protocols
    protocols = conn.execute(
        "SELECT name, dose, frequency, notes, start_date, end_date, track FROM protocols WHERE user_id=%s ORDER BY sort_order, id",
        (uid,)
    ).fetchall()

    # Inventory (physical stock)
    inventory = conn.execute(
        """SELECT name, item_type, total_amount, unit, bac_water_ml, remaining, per_dose,
                  opened_date, status, vendor FROM inventory_items
           WHERE user_id=%s ORDER BY name""",
        (uid,)
    ).fetchall()

    # Dose events (execution log, most recent 500)
    doses = conn.execute(
        """SELECT d.taken_at, d.amount, d.unit, d.site, d.notes,
                  p.name AS protocol_name, i.name AS item_name
           FROM dose_events d
           LEFT JOIN protocols p ON p.id = d.protocol_id
           LEFT JOIN inventory_items i ON i.id = d.item_id
           WHERE d.user_id=%s ORDER BY d.taken_at DESC LIMIT 500""",
        (uid,)
    ).fetchall()

    # Exercise bank
    bank = conn.execute(
        "SELECT name, tool, muscles, mult, is_bw FROM exercises ORDER BY name ASC"
    ).fetchall()

    conn.close()

    # Build structured response
    session_list = []
    # Build name -> muscles lookup for inline enrichment
    muscles_map = {row["name"]: row["muscles"] for row in bank}

    for s in sessions:
        exs = s["exercises"] if isinstance(s["exercises"], list) else json.loads(s["exercises"])
        session_list.append({
            "date":          s["date"],
            "bw":            s["bw"],
            "rd":            s["rd"],
            "notes":         s["notes"] or "",
            "sleep_hours":   s["sleep_hours"],
            "deep_sleep_pct": s["deep_sleep_pct"],
            "exercises": [{
                "name":     ex.get("name"),
                "tool":     ex.get("tool"),
                "muscles":  muscles_map.get(ex.get("name")),
                "totalVol": ex.get("totalVol", 0),
                "density":  ex.get("density", 0),
                "sets": [{
                    "load": st.get("rawLoad", 0),
                    "trueLbs": st.get("trueLbs", 0),
                    "reps": st.get("reps", 0),
                    "time": st.get("time", 1.5),
                    "vol":  st.get("vol", 0)
                } for st in ex.get("sets", [])]
            } for ex in exs]
        })

    return {
        "meta": {
            "username":    username,
            "generated":   datetime.utcnow().isoformat() + "Z",
            "total_sessions": len(session_list),
            "date_range":  {
                "from": session_list[0]["date"] if session_list else None,
                "to":   session_list[-1]["date"] if session_list else None
            }
        },
        "profile": {
            "first_name":     prof["first_name"]     if prof else None,
            "last_name":      prof["last_name"]       if prof else None,
            "dob":            prof["dob"]             if prof else None,
            "gender":         prof["gender"]          if prof else None,
            "height_in":      prof["height_in"]       if prof else None,
            "target_bw":      prof["target_bw"]       if prof else None,
            "activity_level": prof["activity_level"]  if prof else None,
            "week_start":     prof["week_start"]       if prof else "Saturday",
        } if prof else {},
        "phases":    [dict(p) for p in phases],
        "protocols": [dict(p) for p in protocols],
        "inventory": [dict(i) for i in inventory],
        "dose_events": [dict(d) for d in doses],
        "exercise_bank": [dict(b) for b in bank],
        "sessions":  session_list
    }
